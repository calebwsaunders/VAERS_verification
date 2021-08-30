import pyinputplus as pyip
from datetime import date
import calendar
import openpyxl
import glob
import csv


def get_current_date():
    """Get current date with goal format of: dd MMM YYYY"""
    date_year = date.today().year
    date_month = date.today().month
    month_abbr = calendar.month_abbr[date_month]
    date_day = date.today().day
    return f'{date_day} {month_abbr} {date_year}'

def get_user_input(message):
    """Get input from the user with an individualized message and return the user's input."""
    output = ""
    while True:
        output = input(message)
        print(f"You entered {output}; is this correct?")
        verify = pyip.inputMenu(["Yes", "No"], numbered=True)
        if verify == "Yes":
            break
    return output

def choose_excel_file():
    """Showing the user all of the Excel files in the current working directory and asking them to select one to
    load if they have an ongoing file they are adding to."""
    excel_files_in_directory = glob.glob('*.xlsx')
    print("The following Excel workbooks are in this folder:")
    i = 1
    for file in excel_files_in_directory:
        print(f"{i}: {file}")
        i += 1
    load_current_file = pyip.inputMenu(['Yes', 'No'],
                                       "\nDo you want to pick one of these files to load for the output file?\n",
                                       numbered=True)
    if load_current_file == 'Yes':
        output = pyip.inputMenu(excel_files_in_directory, numbered=True)
        return output
    else:
        return 'None'

def choose_file(message):
    """Asking the user to clarify which csv file correlates to VAX ID and which to VAX Reports data."""
    files = glob.glob('*.csv')
    print(message)
    output = pyip.inputMenu(files, numbered=True)
    return output

# Variables for counting
total_occurrences = 0
total_deaths = 0
total_er_visits = 0
total_hospitalizations = 0
total_covid_vax_occurrences = 0
total_covid_vax_deaths = 0
total_covid_vax_er_visits = 0
total_covid_vax_hospitalizations = 0

# VAX file structure: Column 0 - VAERS_ID, Column 1 - VAX_TYPE, Column 2 - VAX_MANU,
# Column 3 - VAX_LOT, Column 4 - VAX_DOSE_SERIES, Column 5 - VAX_ROUTE, Column 6 - VAX_SITE,
# Column 7 - VAX_NAME
VAX_file = choose_file('Which file has the vaccine ID information (Ex: VAERSVAX)?')
DATA_file = choose_file("Which file has the vaccine report data (Ex: VAERSDATA)?")

# Choose and read into a list the VAX data.
vax_data = []
with open(VAX_file, 'r', encoding='windows-1252') as file:
    reader = csv.reader(file, delimiter=',')
    headers = next(reader)
    for row in reader:
        vax_data.append(row)

# Setting up a dictionary to read all the VAX data into.
# Key is VAX_NAME, value is a list of VAERS_ID
vax_data_initial = {}

vax_count_variable = 0
while vax_count_variable < len(vax_data):
    vax_name = vax_data[vax_count_variable][7]
    vax_id = vax_data[vax_count_variable][0]
    if vax_name in vax_data_initial:
        vax_data_initial[vax_name].append(vax_id)
    else:
        vax_data_initial[vax_name] = [vax_id]
    vax_count_variable += 1

# Setup a dictionary for each VAERS_ID entry.
# Determining whether the report is due to death.
vax_reports = {}

# DATA file structure:
# Column 0 - VAERS_ID
# Column 1 - RECVDATE
# Column 2 - STATE
# Column 3 - AGE_YRS
# Column 4 - CAGE_YR
# Column 5 - CAGE_MO
# Column 6 - SEX
# Column 7 - RPT_DATE
# Column 8 - SYMPTOM_TEXT
# Column 9 - DIED
# Column 10 - DATEDIED
# Column 11 - L_THREAT
# Column 12 - ER_VISIT
# Column 13 - HOSPITAL
# Column 14 - HOSPDAYS
# Column 15 - X_STAY
# Column 16 - DISABLE
# Column 17 - RECOVD
# Column 18 - VAX_DATE
# Column 19 - ONSET_DATE
# Column 20 - NUMDAYS
vax_data_data = []
with open(DATA_file, 'r', encoding='windows-1252') as file:
    reader = csv.reader(file, delimiter=',')
    headers = next(reader)
    for row in reader:
        vax_data_data.append(row)

data_count_variable = 0
while data_count_variable < len(vax_data_data):
    vaers_id = vax_data_data[data_count_variable][0]
    reported_death = 0
    reported_er_visit = 0
    reported_hospitalization = 0
    if vax_data_data[data_count_variable][9] == "Y":
        reported_death += 1
    if vax_data_data[data_count_variable][12] == "Y":
        reported_er_visit += 1
    if vax_data_data[data_count_variable][13] == "Y":
        reported_hospitalization += 1

    # Add VAERS_ID to dictionary.
    vax_reports[vaers_id] = [reported_death, reported_er_visit, reported_hospitalization]
    data_count_variable += 1

vax_data_by_type = []
for vaccine_type in vax_data_initial:
    vaccine_name = vaccine_type
    total_reported_occurrences = 0
    total_reported_deaths = 0
    total_reported_er_visits = 0
    total_reported_hospitalizations = 0
    for report_id in vax_data_initial[vaccine_type]:
        total_reported_occurrences += 1
        # 0 - reported_death, 1 - reported_er_visit, 2 - reported_hospitalization
        total_reported_deaths += vax_reports[report_id][0]
        total_reported_er_visits += vax_reports[report_id][1]
        total_reported_hospitalizations += vax_reports[report_id][2]

    # Add parsed data to list.
    vax_data_by_type.append([vaccine_name,  # 0
                          total_reported_occurrences,  # 1
                          total_reported_deaths,  # 2
                          total_reported_er_visits,  # 3
                          total_reported_hospitalizations])  # 4

    # Update totals.
    total_occurrences += total_reported_occurrences
    total_deaths += total_reported_deaths
    total_er_visits += total_reported_er_visits
    total_hospitalizations += total_reported_hospitalizations

    # Update COVID19 vaccine totals.
    if vaccine_type.__contains__('COVID19'):
        total_covid_vax_occurrences += total_reported_occurrences
        total_covid_vax_deaths += total_reported_deaths
        total_covid_vax_er_visits += total_reported_er_visits
        total_covid_vax_hospitalizations += total_reported_hospitalizations

sorted_vax_data_list = sorted(vax_data_by_type, key=lambda vax_deaths: vax_deaths[2], reverse=True)

# A variable for the date of the current data.
data_date = get_user_input("What's the date for this data (it's in the name of the zip folder)? ")

# Check to see if output Excel already exists.
# Load sheet if exists, else create new file.
chosen_file = choose_excel_file()
output_wb = ""
if chosen_file == 'None':
    output_wb = openpyxl.Workbook()
    chosen_file = get_user_input("What would you like to name the file? ")
else:
    output_wb = openpyxl.load_workbook(chosen_file)

output_wb_sheet = output_wb.create_sheet(index=0, title=data_date)
output_wb_sheet.merge_cells('A1:D1')
output_wb_sheet['A1'] = f"VAERS Data from: {data_date}; Parsed  on: {get_current_date()}"
output_wb_sheet['A2'] = "Vaccine Type"
output_wb_sheet['B2'] = "Number of Reports"
output_wb_sheet['C2'] = "Deaths Reported"
output_wb_sheet['D2'] = "ER Visits Reported"
output_wb_sheet['E2'] = "Hospitalizations Reported"

row_to_write_to = 3  # Starting at 3 since the date is going in 1 and headers in 2.
for vaccine in sorted_vax_data_list:
    # Write values to Excel.
    output_wb_sheet[f'A{row_to_write_to}'] = vaccine[0]
    output_wb_sheet[f'B{row_to_write_to}'] = vaccine[1]
    output_wb_sheet[f'C{row_to_write_to}'] = vaccine[2]
    output_wb_sheet[f'D{row_to_write_to}'] = vaccine[3]
    output_wb_sheet[f'E{row_to_write_to}'] = vaccine[4]
    row_to_write_to += 1

# Writing out the totals and comparing COVID19 to everything else.
output_wb_sheet['G2'] = "Total Deaths"
output_wb_sheet['G3'] = total_deaths
output_wb_sheet['G5'] = "COVID19 Vaccine Deaths"
output_wb_sheet['G6'] = total_covid_vax_deaths
output_wb_sheet['G8'] = "Non-COVID Vaccine Deaths"
output_wb_sheet['G9'] = total_deaths - total_covid_vax_deaths
output_wb_sheet['G11'] = "Total ER Visits"
output_wb_sheet['G12'] = total_er_visits
output_wb_sheet['G14'] = "COVID19 ER Visits"
output_wb_sheet['G15'] = total_covid_vax_er_visits
output_wb_sheet['G17'] = "Non-COVID ER Visits"
output_wb_sheet['G18'] = total_er_visits - total_covid_vax_er_visits
output_wb_sheet['G20'] = "Total Hospitalizations"
output_wb_sheet['G21'] = total_hospitalizations
output_wb_sheet['G23'] = "COVID19 Hospitalizations"
output_wb_sheet['G24'] = total_covid_vax_hospitalizations
output_wb_sheet['G26'] = "Non-COVID Hospitalizations"
output_wb_sheet['G27'] = total_hospitalizations - total_covid_vax_hospitalizations
output_wb_sheet['I2'] = "Total Reports"
output_wb_sheet['I3'] = total_occurrences
output_wb_sheet['I5'] = "COVID19 Reports"
output_wb_sheet['I6'] = total_covid_vax_occurrences
output_wb_sheet['I8'] = "Non-COVID Reports"
output_wb_sheet['I9'] = total_occurrences - total_covid_vax_occurrences

# Clean up the spreadsheet.
sheets = output_wb.sheetnames
if 'Sheet' in sheets:
    del output_wb['Sheet']

if chosen_file.endswith('.xlsx'):
    output_wb.save(chosen_file)
    output_wb.close()
else:
    output_wb.save(f'{chosen_file}.xlsx')
    output_wb.close()